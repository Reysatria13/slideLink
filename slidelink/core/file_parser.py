"""
SlideLink Universal File Parser

Extracts text content from various file formats (PDF, DOCX, CSV, Excel, TXT)
to be used as raw data for Slide Generation.
"""

import os
import json
import csv
from dataclasses import dataclass

@dataclass
class ParsedContent:
    text: str
    file_type: str
    metadata: dict


def parse_pdf(file_path: str) -> str:
    try:
        import PyPDF2
        text = ""
        with open(file_path, "rb") as f:
            reader = PyPDF2.PdfReader(f)
            for page in reader.pages:
                text += page.extract_text() + "\\n\\n"
        return text
    except Exception as e:
        return f"[PDF Parsing Error: {e}]"


def parse_docx(file_path: str) -> str:
    try:
        from docx import Document
        doc = Document(file_path)
        return "\\n".join([p.text for p in doc.paragraphs])
    except Exception as e:
        return f"[DOCX Parsing Error: {e}]"


def parse_csv(file_path: str) -> str:
    try:
        text = ""
        with open(file_path, newline='', encoding='utf-8') as f:
            reader = csv.reader(f)
            for row in reader:
                text += " | ".join(row) + "\\n"
        return text
    except Exception as e:
        return f"[CSV Parsing Error: {e}]"


def parse_excel(file_path: str) -> str:
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file_path, data_only=True)
        text = ""
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            text += f"--- Sheet: {sheet_name} ---\\n"
            for row in sheet.iter_rows(values_only=True):
                text += " | ".join([str(cell) if cell is not None else "" for cell in row]) + "\\n"
            text += "\\n"
        return text
    except Exception as e:
        return f"[Excel Parsing Error: {e}]"


def parse_txt(file_path: str) -> str:
    try:
        with open(file_path, "r", encoding='utf-8') as f:
            return f.read()
    except Exception as e:
        return f"[TXT Parsing Error: {e}]"


def parse_file(file_path: str, filename: str) -> ParsedContent:
    """
    Identifies the file type and routes it to the correct parser.
    """
    ext = os.path.splitext(filename)[1].lower()
    text = ""
    file_type = ext.replace(".", "")

    if ext == ".pdf":
        text = parse_pdf(file_path)
    elif ext in [".docx", ".doc"]:
        text = parse_docx(file_path)
    elif ext == ".csv":
        text = parse_csv(file_path)
    elif ext in [".xlsx", ".xls"]:
        text = parse_excel(file_path)
    elif ext in [".txt", ".md", ".json"]:
        text = parse_txt(file_path)
    else:
        text = f"[Unsupported file type: {ext}]. Cannot parse raw content directly."

    # Limit extremely large files to prevent massive token usage
    max_chars = 100000 
    if len(text) > max_chars:
        text = text[:max_chars] + "\\n\\n...[Content truncated due to size limit]..."

    return ParsedContent(
        text=text.strip(),
        file_type=file_type,
        metadata={"filename": filename, "chars": len(text)}
    )
